### IEC61400-15-1 site suitability input form converter (from *.xlsx to *.json)
## Usage:
## xlsx2json.py [Input xlsx file] [Output json file]

import openpyxl
import json
import sys

#
args = sys.argv
InputXlsxFile = args[1]
OutputJsonFile = args[2]

wb = openpyxl.load_workbook(InputXlsxFile,data_only=True)
ws = wb['Project Information']
project_name = ws.cell(1,2).value
project_owner = ws.cell(2,2).value
project_number = ws.cell(3,2).value
author_name = ws.cell(4,2).value
project_date = str(ws.cell(5,2).value)
revision_number = ws.cell(6,2).value
country = ws.cell(7,2).value
datum = ws.cell(8,2).value
projection = ws.cell(9,2).value
report_filename = ws.cell(10,2).value
report_revision_number = ws.cell(11,2).value
def_version = ws.cell(12,2).value

print("Analyzing the excel file...")

# Analyze the wind speed bin width (wsbw) and number of wind direction sectors (nwds)
ws = wb['WS Frequency']
wsbw = int(ws.cell(5,2).value)
nwds = int(360 / ws.cell(int (40 / wsbw + 5),1).value)

print("The wind speed bin width is",wsbw)
print("The number of wind direction sector is",nwds)

nwsb = int(40 / wsbw) + 1


# Analyze the number of wind turbines and their IDs
ws = wb['Turbine Layout Summary']
wtids = []
row = 2
while ws.cell(row,2).value != None:
    wtids.append(str(ws.cell(row,2).value))
    row = row + 1

nwt = len(wtids)

print("The number of wind turbine is",nwt)
print("Their IDs are",wtids)

# Analyse the number of measurement devices and their IDs
ws = wb['Measurement Device Summary']
mdids = []
row = 2
while ws.cell(row,1).value != None:
    mdids.append(str(ws.cell(row,1).value))
    row = row + 1

nmd = len(mdids)

print("The number of measurement devices is",nmd)
print("Their IDs are",mdids)

# Make the whole list of WT + meas. device
mdwtids = mdids + wtids


# Meta data
OjMetaData = {}
OjMetaData["Number of wind direction secttors"] = nwds
OjMetaData["Wind speed bin width"] = wsbw
OjMetaData["Number of measurement devices"] = nmd
OjMetaData["Measurement device IDs"] = mdids
OjMetaData["Number of wind turbines"] = nwt
OjMetaData["Wind turbine IDs"] = wtids

# Project information
OjPI = {}
OjPI["Name"] = project_name
OjPI["Owner"] = project_owner
OjPI["Number"] = project_number
OjPI["Author"] = author_name
OjPI["Date"] = project_date
OjPI["Revision number and reason"] = revision_number
OjPI["Country"] = country
OjPI["Datum"] = datum
OjPI["Projection"] = projection
OjPI["Report filename"] = report_filename
OjPI["Report revision number"] = report_revision_number
OjPI["Dgital exchange format version"] = def_version

# Turbine Layout Summary
OjTLS = {}
ws = wb['Turbine Layout Summary']
for wt in wtids:
    row = 2
    while (row < nwt + 2):
        if (str(ws.cell(row,2).value)) == wt:
            trow = row
            exit
        row = row + 1

    tmpobj = {}
    tmpobj["Project Name"] = ws.cell(trow,1).value
    tmpobj["Easting or Latitude"] = ws.cell(trow,3).value
    tmpobj["Northing or Latitude"] = ws.cell(trow,4).value
    tmpobj["Ground Elevation"] = ws.cell(trow,5).value
    tmpobj["Wind Turbine Manufacturer"] = ws.cell(trow,6).value
    tmpobj["Model"] = ws.cell(trow,7).value
    tmpobj["Rated Power"] = ws.cell(trow,8).value
    tmpobj["Rotor Diameter"] = ws.cell(trow,9).value
    tmpobj["Hub Height"] = ws.cell(trow,10).value
    tmpobj["Data Source"] = ws.cell(trow,11).value
    tmpobj["Ve50"] = ws.cell(trow,12).value
    tmpobj["V50"] = ws.cell(trow,13).value
    tmpobj["COV"] = ws.cell(trow,14).value
    tmpobj["Air Density"] = ws.cell(trow,15).value
    tmpobj["Annual Average Wind Speed"] = ws.cell(trow,16).value
    tmpobj["Weibull Scale Parameter"] = ws.cell(trow,17).value
    tmpobj["Weibull Shape Parameter "] = ws.cell(trow,18).value
    tmpobj["CCT"] = ws.cell(trow,19).value
    tmpobj["Annual Mean Wind Shear"] = ws.cell(trow,20).value
    tmpobj["TI15"] = ws.cell(trow,21).value
    tmpobj["Sigma I"] = ws.cell(trow,22).value
    tmpobj["Inflow Angle"] = ws.cell(trow,23).value

    OjTLS[wt] = tmpobj

# Measurement Device Summary
OjMDS = {}
ws = wb['Measurement Device Summary']
for md in mdids:
    row = 2
    while (row <= nmd + 2):
        if (str(ws.cell(row,1).value)) == md:
            trow = row
            exit
        row = row + 1

    EachMDS = {}
    EachMDS["Easting or Latitude"] = ws.cell(trow,2).value
    EachMDS["Northing or Latitude"] = ws.cell(trow,3).value
    EachMDS["Ground Elevation"] = ws.cell(trow,4).value
    EachMDS["Measurement Device Height"] = ws.cell(trow,5).value

    OjMDS[md] = EachMDS


# WS Frequency
OjWSF = {}
ws = wb['WS Frequency']

# For measurement device
for md in mdids:
    column = 3
    while (column <= nwt + nmd*2 + 3):
        if (str(ws.cell(3,column).value)) == md:
            tcolumn = column
            exit
        column = column + 1

    tmpobj = {}

    # WS Frequency
    tmparray2 = []
    iwd = 1
    while iwd <= nwds:
        tmparray = []
        iws = 0
        while iws <= 40:
            row = 4+(iwd-1)*nwsb + iws
            tmparray.append(ws.cell(row,tcolumn).value * 100.0)
            iws = iws + wsbw
        tmparray2.append(tmparray)
        iwd = iwd + 1

    tmpobj['WS Frequency'] = tmparray2

    # WS number of samples
    tmparray2 = []
    iwd = 1
    while iwd <= nwds:
        tmparray = []
        iws = 0
        while iws <= 40:
            row = 4+(iwd-1)*nwsb + iws
            tmparray.append(int(ws.cell(row,tcolumn + 1).value))
            iws = iws + wsbw
        tmparray2.append(tmparray)
        iwd = iwd + 1

    tmpobj['WS Number of samples'] = tmparray2

    OjWSF[md]=tmpobj


# For wind turbines
for wt in wtids:
    column = 3
    while (column <= nwt + nmd*2 + 3):
        if (str(ws.cell(3,column).value)) == wt:
            tcolumn = column
            exit
        column = column + 1

    tmpobj = {}
    # WS Frequency
    tmparray2 = []
    iwd = 1
    while iwd <= nwds:
        tmparray = []
        iws = 0
        while iws <= 40:
            row = 4+(iwd-1)*nwsb + iws
            tmparray.append(ws.cell(row,tcolumn).value * 100.0)
            iws = iws + wsbw
        tmparray2.append(tmparray)
        iwd = iwd + 1

    tmpobj['WS Frequency'] = tmparray2

    OjWSF[wt]=tmpobj


# WS Weibull
ws = wb['WS Weibull']
OjWSW = {}

# For both meas. device and wt
for mdwt in mdwtids:
    column = 3
    tmpobj={}
    while (column <= nwt + nmd + 2):
        if (str(ws.cell(3,column).value)) == mdwt:
            tcolumn = column
            exit
        column = column + 1

    tmpobj['WS Weibull scale parameter all directions'] = ws.cell(4,tcolumn).value
    tmpobj['WS Weibull shape parameter all directions'] = ws.cell(5 + nwds,tcolumn).value

    tmparray1=[]
    tmparray2=[]
    tmparray3=[]
    iwd = 1
    while(iwd <= nwds):
        tmparray1.append(ws.cell(3 + iwd,tcolumn).value)
        tmparray2.append(ws.cell(5 + nwds + iwd,tcolumn).value)
        tmparray3.append(ws.cell(5 + 2 * nwds + iwd, tcolumn).value * 100.0)
        iwd = iwd + 1

    tmpobj['WS Weibull scale parameter'] = tmparray1
    tmpobj['WS Weibull shape parameter'] = tmparray2
    tmpobj['WS Weibull frequency'] = tmparray3

    OjWSW[mdwt]=tmpobj


## Ambient Mean TI
ws = wb['Ambient Mean TI']
OjAMTI = {}

# For meas. device and WT
for mdwt in mdwtids:
    column = 3
    while (column <= nwt + nmd + 2):
        if (str(ws.cell(3,column).value)) == mdwt:
            tcolumn = column
            exit
        column = column + 1

    tmpobj = {}

    # For all wind direction
    tmparray = []
    iws = 0
    while iws <= 40:
        row = 4 + iws
        tmparray.append(ws.cell(row,tcolumn).value*100.0)
        iws = iws + wsbw
    tmpobj["Ambient mean TI all directions"] = tmparray


    # For each wind direction
    tmparray2 = []
    iwd = 1
    while iwd <= nwds:
        tmparray = []
        iws = 0
        while iws <= 40:
            row = 4 + iwd * nwsb + iws
            tmparray.append(ws.cell(row,tcolumn).value * 100.0)
            iws = iws + wsbw
        tmparray2.append(tmparray)
        iwd = iwd + 1
    tmpobj['Ambient mean TI'] = tmparray2


    OjAMTI[mdwt] = tmpobj


## SD TI
ws = wb['SD TI']
OjSDTI = {}

# For meas. device and WT
for mdwt in mdwtids:
    column = 3
    while (column <= nwt + nmd + 2):
        if (str(ws.cell(3,column).value)) == mdwt:
            tcolumn = column
            exit
        column = column + 1

    tmpobj = {}

    # For all wind direction
    tmparray = []
    iws = 0
    while iws <= 40:
        row = 4 + iws
        tmparray.append(ws.cell(row,tcolumn).value*100.0)
        iws = iws + wsbw
    tmpobj["SD TI all directions"] = tmparray


    # For each wind direction
    tmparray2 = []
    iwd = 1
    while iwd <= nwds:
        tmparray = []
        iws = 0
        while iws <= 40:
            row = 4 + iwd * nwsb + iws
            tmparray.append(ws.cell(row,tcolumn).value * 100.0)
            iws = iws + wsbw
        tmparray2.append(tmparray)
        iwd = iwd + 1
    tmpobj['SD TI'] = tmparray2

    OjSDTI[mdwt] = tmpobj

## Extreme Ambient TI
ws = wb['Extreme Ambient TI']
OjEATI = {}

# For meas. device and WT
for mdwt in mdwtids:
    column = 3
    while (column <= nwt + nmd + 2):
        if (str(ws.cell(3,column).value)) == mdwt:
            tcolumn = column
            exit
        column = column + 1

    tmpobj = {}

    # For all wind direction
    tmparray = []
    iws = 0
    while iws <= 40:
        row = 4 + iws
        tmparray.append(ws.cell(row,tcolumn).value * 100)
        iws = iws + wsbw
    tmpobj["SD TI all directions"] = tmparray

    OjEATI[mdwt]= tmpobj

## Temperature
ws = wb['Temperature']
OjTEMP = {}

# For meas. device and WT
for mdwt in mdwtids:
    ## Prepare empty dictionary
    tmpobj = {}

    ## Scalar variables
    column = 2
    while (column <= 2 * nwt + 2 * nmd + 1):
        if (str(ws.cell(3,column).value)) == mdwt:
            tcolumn = column
            exit
        column = column + 1

    tmpobj['Yearly mean ambient Temperature'] = ws.cell(4,tcolumn).value
    tmpobj['Days per year with at least 1 hour below -20 deg'] = ws.cell(5,tcolumn).value

    ## Vector variables
    column = 2
    while (column <= 2 * nwt + 2* nmd + 1):
        if (str(ws.cell(3,column).value)) == mdwt:
            tcolumn = column
            exit
        column = column + 1

    tmparray1 = []
    tmparray2 = []
    itemp = 1
    while itemp <= 91:
        row = 8 + itemp
        tmparray1.append(ws.cell(row,tcolumn).value)
        tmparray2.append(int(ws.cell(row,tcolumn+1).value))
        itemp = itemp + 1
    tmpobj["Temperature frequency"] = tmparray1
    tmpobj["Number of samples"] = tmparray2

    OjTEMP[mdwt] = tmpobj

## Shear
ws = wb['Shear']
OjSHEAR = {}

# For meas. device and WT
for mdwt in mdwtids:
    tmpobj = {}
    column = 3
    while (column <= nwt + nmd + 2):
        if (str(ws.cell(3,column).value)) == mdwt:
            tcolumn = column
            exit
        column = column + 1

    tmpobj["Shear all directions"] = ws.cell(4,tcolumn).value
    tmparray = []
    iwd = 1
    while iwd <= nwds:
        row = 4 + iwd
        tmparray.append(ws.cell(row,tcolumn).value)
        iwd = iwd + 1

    tmpobj["Directional shear"] = tmparray

    OjSHEAR[mdwt]=tmpobj


## Inflow Angle
ws = wb['Inflow Angle']
OjIA = {}

# For meas. device and WT
for mdwt in mdwtids:
    tmpobj = {}
    column = 2
    while (column <= nwt + nmd + 1):
        if (str(ws.cell(3,column).value)) == mdwt:
            tcolumn = column
            exit
        column = column + 1

    tmpobj["Inflow angle all directions"] = ws.cell(4,tcolumn).value
    tmpobj["Inflow angle max"] = ws.cell(5,tcolumn).value

    tmparray = []
    iwd = 1
    while iwd <= nwds:
        row = 5 + iwd
        tmparray.append(ws.cell(row,tcolumn).value)
        iwd = iwd + 1

    tmpobj["Directional Inflow angle"] = tmparray

    OjIA[mdwt]=tmpobj

##CcT
ws = wb['CcT']
OjCCT = {}

# For meas. device and WT
for mdwt in mdwtids:
    tmpobj = {}
    column = 2
    while (column <= nwt + nmd + 1):
        if (str(ws.cell(3,column).value)) == mdwt:
            tcolumn = column
            exit
        column = column + 1

    tmpobj["sigma 3/sigma 1"] = ws.cell(4,tcolumn).value
    tmpobj["sigma 2/sigma 1"] = ws.cell(5,tcolumn).value
    tmpobj["CcT"] = ws.cell(6,tcolumn).value

    OjCCT[mdwt]=tmpobj


# Make top level keys
top = {}
top["Project Information"] = OjPI
top["Meta data"] = OjMetaData
top["Turbine Layout Summary"] = OjTLS
top["Measurement Device Summary"] = OjMDS
top["WS frequency"] = OjWSF
top["WS Weibull"] = OjWSW
top["Ambient Mean TI"] = OjAMTI
top["SD TI"] = OjSDTI
top["Extreme Ambient TI"] = OjEATI
top["Temperature"] = OjTEMP
top["Shear"] = OjSHEAR
top["Inflow Angle"] = OjIA
top["CcT"] = OjCCT


# Damp the json file
with open(OutputJsonFile, 'w') as f:
    json.dump(top, f, ensure_ascii=False, indent=4)
